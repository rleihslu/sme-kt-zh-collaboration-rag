from loguru import logger
import openpyxl  # type: ignore[import-untyped]

from conversational_toolkit.chunking.base import Chunk, Chunker


class ExcelChunker(Chunker):
    """
    Chunks Excel workbooks (.xlsx, .xls) into one Chunk per sheet.

    Each sheet is converted to a Markdown table where the first row is treated
    as the header. Empty sheets are skipped. Metadata includes the sheet name,
    row count, and column count.
    """

    def make_chunks(self, file_path: str) -> list[Chunk]:
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            logger.error(f"Error reading {file_path}: {e}")
            return []

        chunks: list[Chunk] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = [
                [cell.value for cell in row] for row in ws.iter_rows() if any(cell.value is not None for cell in row)
            ]

            if not rows:
                continue

            header = [str(cell) if cell is not None else "" for cell in rows[0]]
            separator = ["---"] * len(header)

            lines = [
                "| " + " | ".join(header) + " |",
                "| " + " | ".join(separator) + " |",
            ]
            for row in rows[1:]:
                cells = [str(cell) if cell is not None else "" for cell in row]
                # Pad shorter rows to match header length
                while len(cells) < len(header):
                    cells.append("")
                lines.append("| " + " | ".join(cells) + " |")

            chunks.append(
                Chunk(
                    title=sheet_name,
                    content="\n".join(lines),
                    mime_type="text/markdown",
                    metadata={
                        "sheet": sheet_name,
                        "rows": len(rows) - 1,
                        "columns": len(header),
                    },
                )
            )

        return chunks
